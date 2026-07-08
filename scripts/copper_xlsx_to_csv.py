"""Copper full-export .xlsx -> import-ready .csv.
Keeps only columns the CRM importer consumes; drops the 200+ padded
Website/Social slots and Copper-internal ids."""
import csv
import re
import sys
from datetime import datetime, date
from pathlib import Path

from openpyxl import load_workbook

NAMED = {
    "people": {"First Name", "Middle Name", "Last Name", "Prefix", "Suffix", "Title",
               "Details", "Company", "Owned By", "Street", "City", "State", "Postal Code",
               "Country", "Contact Type", "Last Contacted", "Created At"},
    "leads": {"First Name", "Middle Name", "Last Name", "Prefix", "Suffix", "Title",
              "Details", "Value", "Account", "Owned By", "Status", "Source", "Street",
              "City", "State", "Postal Code", "Country", "Converted At", "Created At"},
    "companies": {"Name", "Details", "Email Domain", "Street", "City", "State",
                  "Postal Code", "Country", "Owned By", "Contact Type", "Created At"},
    "opportunities": {"Name", "Details", "Company", "Primary Person Contact", "Status",
                      "Priority", "Owner", "Close Date", "Value", "Win Probability",
                      "Pipeline", "Stage", "Source", "Loss Reason", "Created At"},
}
PAIR = re.compile(r"^(Email|Phone Number|Social|Website)( (\d+))?( Type)?$")
CF = re.compile(r"^.+\s+cf_\d+(\s+people\s+(names|ids))?$")
PAIR_LIMIT = {"Email": 4, "Phone Number": 9, "Social": 6, "Website": 2}


def keep(h: str, kind: str) -> bool:
    if h in NAMED[kind] or h in ("Tag", "Tags"):
        return True
    if CF.match(h):
        return not h.endswith(" ids")
    m = PAIR.match(h)
    if m:
        return int(m.group(3) or 1) <= PAIR_LIMIT[m.group(1)]
    return False


def cell_to_str(v):
    if v is None:
        return ""
    if isinstance(v, datetime):
        if v.hour == v.minute == v.second == 0:
            return f"{v.month}/{v.day}/{v.year}"
        return v.strftime("%m/%d/%Y %H:%M")
    if isinstance(v, date):
        return f"{v.month}/{v.day}/{v.year}"
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


downloads = Path.home() / "Downloads"
for kind in ("people", "leads", "companies", "opportunities"):
    src = downloads / f"{kind}.xlsx"
    if not src.exists():
        print(f"{kind}: no xlsx, skipped")
        continue
    wb = load_workbook(src, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    headers = [cell_to_str(c).strip() for c in next(it)]
    cols = [i for i, h in enumerate(headers) if keep(h, kind)]
    out = downloads / f"{kind}.csv"
    n = 0
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([headers[i] for i in cols])
        for row in it:
            vals = [cell_to_str(row[i]) if i < len(row) else "" for i in cols]
            if any(v.strip() for v in vals):
                w.writerow(vals)
                n += 1
    wb.close()
    print(f"{kind}: {n} rows, {len(headers)} -> {len(cols)} columns -> {out.name}")
